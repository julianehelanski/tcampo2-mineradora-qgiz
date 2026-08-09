# -*- coding: utf-8 -*-
"""
Análise temporal dos processos minerários do SIGMINE/ANM (MS).

Este script NÃO roda no QGIS — é Python comum (pandas + openpyxl + matplotlib).
Ele lê a planilha SIGMINE_MS_processos_minerarios_BellaPedra.xlsx e:

1. Extrai as informações temporais escondidas em colunas de texto:
   - Ano de abertura  → parte final do nº do processo (NNNNNN/AAAA);
   - Data do último evento → trecho "EM DD/MM/AAAA" da coluna "Último evento".
2. Grava uma aba nova "Dados temporais" com as colunas extraídas.
3. Grava uma aba "Séries e gráficos" com tabelas agregadas e gráficos
   nativos do Excel (série temporal e por substância).
4. Exporta os mesmos gráficos como PNG na pasta graficos/.

Uso (na raiz do repositório):
    python scripts/analise_temporal_sigmine.py [caminho_da_planilha.xlsx]

Limitação da fonte: o SIGMINE informa apenas o ÚLTIMO evento de cada
processo, não o histórico completo. O ano no nº do processo é o ano do
PROTOCOLO do pedido na ANM, não do início da extração real.
"""

import re
import sys
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------- parâmetros

RAIZ = Path(__file__).resolve().parent.parent
ENTRADA_PADRAO = RAIZ / "dados" / "reais" / "SIGMINE_MS_BellaPedra_temporal.xlsx"
PASTA_GRAFICOS = RAIZ / "graficos"

ABA_DADOS = "Todos os processos (MS)"
ABA_NOVA = "Dados temporais"
ABA_GRAFICOS = "Séries e gráficos"

# Data de referência para calcular idade e tempo sem movimentação.
DATA_REF = date(2026, 8, 9)

# Área aproximada de MS (357.145 km²), para as marcas de "% do estado".
MS_HA = 35_714_500

# Recorte da série por substância (antes de 2000 os dados são esparsos).
ANO_CORTE = 2000

# Convenções visuais do arquivo original (mantidas nas abas novas)
AZUL_ESCURO = "1F4E79"   # cabeçalhos
ROSA_BELLA = "F4CCE0"    # linhas da Bella Pedra
FONTE = "Arial"

# Paleta categórica dos gráficos (uma cor fixa por substância)
CORES_SERIES = ["2A78D6", "EB6834", "1BAF7A", "EDA100", "E87BA4"]
COR_TOTAL = "256ABF"     # série agregada (todos os processos)
COR_CALCARIO = "008300"  # grupo do calcário (foco do trabalho de campo)

# Bolhas: TODAS as substâncias, agrupadas em famílias — cada família com
# uma cor. A classificação é nossa (não vem da ANM); ajuste à vontade.
FAMILIAS = [
    ("CONSTRUÇÃO CIVIL", "#2a78d6", [
        "AREIA", "BASALTO", "CASCALHO", "ARGILA", "ARENITO", "SAIBRO",
        "GRANITO", "QUARTZITO", "BASALTO P/ BRITA", "DIABÁSIO",
        "ARGILA REFRATÁRIA", "ARGILA VERMELHA", "ARDÓSIA", "CONGLOMERADO"]),
    ("METÁLICOS", "#eb6834", [
        "MINÉRIO DE FERRO", "MINÉRIO DE COBRE", "ILMENITA",
        "MINÉRIO DE OURO", "OURO", "MINÉRIO DE MANGANÊS", "FERRO",
        "MANGANÊS", "HEMATITA", "MINÉRIO DE ZINCO", "MINÉRIO DE CHUMBO",
        "MINÉRIO DE ARSÊNICO", "MINÉRIO DE TITÂNIO", "MOLIBDENITA"]),
    ("CALCÁRIO E CARBONATOS", "#008300", [
        "CALCÁRIO", "CALCÁRIO CALCÍTICO", "CALCÁRIO DOLOMÍTICO",
        "CALCITA", "MÁRMORE", "DOLOMITO"]),
    ("INDUSTRIAIS", "#4a3aa7", [
        "FOSFATO", "TURFA", "GRAFITA", "FOLHELHO"]),
    ("GEMAS E COLEÇÃO", "#e87ba4", [
        "DIAMANTE", "QUARTZO", "AMETISTA", "DIAMANTE INDUSTRIAL"]),
    ("ÁGUA MINERAL", "#1baf7a", ["ÁGUA MINERAL", "ÁGUAS TERMAIS"]),
    ("NÃO CADASTRADO", "#c3c2b7", ["DADO NÃO CADASTRADO"]),
]

SUBSTANCIAS_CALCARIO = [
    "CALCÁRIO", "CALCÁRIO CALCÍTICO", "CALCÁRIO DOLOMÍTICO", "CALCITA",
]

# ---------------------------------------------------------------- extração


def extrair_temporal(df: pd.DataFrame) -> pd.DataFrame:
    """Devolve o DataFrame com as colunas temporais extraídas."""
    saida = df[["Processo", "Titular", "Substância", "Fase",
                "Área (ha)", "Uso"]].copy()

    saida["Ano_abertura"] = pd.to_numeric(
        df["Processo"].str.extract(r"/(\d{4})$")[0])

    evento = df["Último evento"].fillna("")
    saida["Cod_evento"] = pd.to_numeric(
        evento.str.extract(r"^(\d+)\s*-")[0], errors="coerce")
    saida["Desc_evento"] = (
        evento.str.replace(r"^\d+\s*-\s*", "", regex=True)
              .str.replace(r"\s*(?:PROTOC\s+)?EM\s+\d{2}/\d{2}/\d{4}\s*$",
                           "", regex=True)
              .str.strip())

    data_txt = evento.str.extract(r"EM (\d{2}/\d{2}/\d{4})")[0]
    saida["Data_ultimo_evento"] = pd.to_datetime(
        data_txt, format="%d/%m/%Y", errors="coerce")
    saida["Ano_ultimo_evento"] = saida["Data_ultimo_evento"].dt.year

    saida["Idade_processo_anos"] = DATA_REF.year - saida["Ano_abertura"]
    saida["Anos_sem_movimentacao"] = (
        (pd.Timestamp(DATA_REF) - saida["Data_ultimo_evento"]).dt.days
        / 365.25).round(1)

    saida["Bella_Pedra"] = df["Titular"].str.contains(
        "BELLA PEDRA", na=False).map({True: "Sim", False: ""})
    return saida


def blocos_familias(dados: pd.DataFrame):
    """Organiza as substâncias em blocos por família, para o gráfico.

    Devolve [(família, cor, [(substância, série anual), ...]), ...],
    famílias em ordem de total de processos (a cinza sempre por último),
    e dentro de cada família as substâncias da maior para a menor.
    """
    contagem = dados["Substância"].value_counts()
    matriz = (dados.pivot_table(index="Substância", columns="Ano_abertura",
                                values="Processo", aggfunc="count",
                                fill_value=0))

    mapeadas = {s for _, _, subs in FAMILIAS for s in subs}
    soltas = sorted(set(contagem.index) - mapeadas)
    familias = list(FAMILIAS)
    if soltas:  # um SIGMINE futuro pode trazer substâncias novas
        print(f"⚠ Substâncias fora da classificação (entram em cinza): "
              f"{soltas}")
        familias.append(("NÃO CLASSIFICADAS", "#c3c2b7", soltas))

    blocos = []
    for nome_fam, cor, subs in familias:
        presentes = sorted((s for s in subs if s in contagem.index),
                           key=lambda s: -contagem[s])
        if presentes:
            blocos.append((nome_fam, cor,
                           [(s, matriz.loc[s]) for s in presentes]))
    cinzas = {"NÃO CADASTRADO", "NÃO CLASSIFICADAS"}
    blocos.sort(key=lambda b: (b[0] in cinzas,
                               -sum(s.sum() for _, s in b[2])))
    return blocos


def resumo_bolhas(dados: pd.DataFrame) -> pd.DataFrame:
    """Tabela-resumo por substância (companheira do gráfico de bolhas)."""
    linhas = []
    for nome_fam, _, itens in blocos_familias(dados):
        for substancia, serie in itens:
            com_valor = serie[serie > 0]
            pico = com_valor.idxmax()
            linhas.append({
                "Família": nome_fam.capitalize(),
                "Substância": substancia,
                "Total de processos": int(serie.sum()),
                "Primeiro ano": int(com_valor.index.min()),
                "Último ano": int(com_valor.index.max()),
                "Ano de pico": int(pico),
                "Processos no pico": int(serie[pico]),
            })
    return pd.DataFrame(linhas)


# ---------------------------------------------------------------- planilha


def _cabecalho(ws, linha, textos, larguras=None):
    fill = PatternFill("solid", fgColor=AZUL_ESCURO)
    for j, texto in enumerate(textos, start=1):
        c = ws.cell(row=linha, column=j, value=texto)
        c.font = Font(name=FONTE, size=9, bold=True, color="FFFFFF")
        c.fill = fill
        c.alignment = Alignment(wrap_text=True, vertical="center")
    if larguras:
        for j, larg in enumerate(larguras, start=1):
            ws.column_dimensions[get_column_letter(j)].width = larg


def _titulo(ws, texto, subtitulo, n_colunas):
    c = ws.cell(row=1, column=1, value=texto)
    c.font = Font(name=FONTE, size=12, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=AZUL_ESCURO)
    for j in range(2, n_colunas + 1):
        ws.cell(row=1, column=j).fill = PatternFill("solid",
                                                    fgColor=AZUL_ESCURO)
    s = ws.cell(row=2, column=1, value=subtitulo)
    s.font = Font(name=FONTE, size=8, color="595959")


def gravar_aba_dados(wb, dados: pd.DataFrame):
    if ABA_NOVA in wb.sheetnames:
        del wb[ABA_NOVA]
    ws = wb.create_sheet(ABA_NOVA)

    _titulo(
        ws,
        "DADOS TEMPORAIS EXTRAÍDOS — processos minerários de MS "
        "(SIGMINE/ANM)",
        "Gerado por scripts/analise_temporal_sigmine.py em "
        f"{DATA_REF.strftime('%d/%m/%Y')} (data de referência dos cálculos "
        "de idade). Ano_abertura = ano do protocolo no nº do processo. "
        "O SIGMINE traz só o ÚLTIMO evento, não o histórico completo. "
        "Linhas em rosa = Bella Pedra Cristal Ltda.",
        len(dados.columns))

    _cabecalho(ws, 4, list(dados.columns),
               larguras=[13, 26, 22, 24, 10, 16, 9, 9, 34, 13, 9, 10, 12, 9])

    rosa = PatternFill("solid", fgColor=ROSA_BELLA)
    fnt = Font(name=FONTE, size=9)
    fnt_bp = Font(name=FONTE, size=9, bold=True)
    for i, linha in enumerate(dados.itertuples(index=False), start=5):
        eh_bp = linha[-1] == "Sim"
        for j, valor in enumerate(linha, start=1):
            if pd.isna(valor):
                valor = None
            elif isinstance(valor, pd.Timestamp):
                valor = valor.date()
            c = ws.cell(row=i, column=j, value=valor)
            c.font = fnt_bp if eh_bp else fnt
            if eh_bp:
                c.fill = rosa
            if dados.columns[j - 1] == "Data_ultimo_evento":
                c.number_format = "DD/MM/YYYY"
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(dados.columns))}" \
                         f"{4 + len(dados)}"


def _estilo_serie(serie, cor, linha=False):
    serie.graphicalProperties.solidFill = cor
    if linha:
        serie.graphicalProperties.line.solidFill = cor
        serie.graphicalProperties.line.width = 20000  # ~2 px
        serie.smooth = False
    else:
        serie.graphicalProperties.line.noFill = True


def gravar_aba_graficos(wb, dados: pd.DataFrame):
    if ABA_GRAFICOS in wb.sheetnames:
        del wb[ABA_GRAFICOS]
    ws = wb.create_sheet(ABA_GRAFICOS)

    _titulo(ws,
            "SÉRIES TEMPORAIS E RECORTES POR SUBSTÂNCIA",
            "Tabelas agregadas a partir da aba 'Dados temporais'. "
            "Os gráficos ao lado são nativos do Excel e podem ser editados. "
            "Versões em PNG estão na pasta graficos/ do repositório.",
            8)

    # ---- T1: aberturas por ano (todos os processos) -----------------------
    anos = range(int(dados["Ano_abertura"].min()),
                 int(dados["Ano_abertura"].max()) + 1)
    por_ano = dados["Ano_abertura"].value_counts().reindex(anos, fill_value=0)

    lin = 4
    _cabecalho(ws, lin, ["Ano", "Processos abertos"], larguras=[10, 16])
    for ano, qtd in por_ano.items():
        lin += 1
        ws.cell(row=lin, column=1, value=ano).font = Font(name=FONTE, size=9)
        ws.cell(row=lin, column=2, value=int(qtd)).font = Font(name=FONTE,
                                                               size=9)
    fim_t1 = lin

    g1 = LineChart()
    g1.title = "Processos minerários abertos por ano — MS (SIGMINE/ANM)"
    g1.y_axis.title = "Processos abertos"
    g1.x_axis.title = "Ano do protocolo"
    g1.height, g1.width = 9, 24
    g1.add_data(Reference(ws, min_col=2, min_row=4, max_row=fim_t1),
                titles_from_data=True)
    g1.set_categories(Reference(ws, min_col=1, min_row=5, max_row=fim_t1))
    _estilo_serie(g1.series[0], COR_TOTAL, linha=True)
    g1.legend = None
    ws.add_chart(g1, "D4")

    # ---- T2: processos por substância (top 10 + outras) -------------------
    contagem = dados["Substância"].value_counts()
    top10 = contagem.head(10)
    outras = int(contagem.iloc[10:].sum())

    lin0 = fim_t1 + 3
    _cabecalho(ws, lin0, ["Substância", "Processos"], larguras=None)
    lin = lin0
    for nome, qtd in top10.items():
        lin += 1
        ws.cell(row=lin, column=1, value=nome).font = Font(name=FONTE, size=9)
        ws.cell(row=lin, column=2, value=int(qtd)).font = Font(name=FONTE,
                                                               size=9)
    lin += 1
    ws.cell(row=lin, column=1, value="OUTRAS").font = Font(name=FONTE, size=9,
                                                           italic=True)
    ws.cell(row=lin, column=2, value=outras).font = Font(name=FONTE, size=9,
                                                         italic=True)
    fim_t2 = lin

    g2 = BarChart()
    g2.type = "col"
    g2.title = "Processos por substância — MS (top 10 + outras)"
    g2.y_axis.title = "Processos"
    g2.height, g2.width = 9, 24
    g2.add_data(Reference(ws, min_col=2, min_row=lin0, max_row=fim_t2),
                titles_from_data=True)
    g2.set_categories(Reference(ws, min_col=1, min_row=lin0 + 1,
                                max_row=fim_t2))
    _estilo_serie(g2.series[0], CORES_SERIES[0])
    g2.legend = None
    g2.gapWidth = 40
    ws.add_chart(g2, f"D{lin0}")

    # ---- T3: série por substância (top 5, a partir de ANO_CORTE) ----------
    top5 = list(contagem.head(5).index)
    recorte = dados[dados["Ano_abertura"] >= ANO_CORTE]
    tabela = (recorte[recorte["Substância"].isin(top5)]
              .pivot_table(index="Ano_abertura", columns="Substância",
                           values="Processo", aggfunc="count", fill_value=0)
              .reindex(range(ANO_CORTE, DATA_REF.year + 1), fill_value=0)
              [top5])

    lin0 = fim_t2 + 3
    _cabecalho(ws, lin0, ["Ano"] + top5)
    lin = lin0
    for ano, valores in tabela.iterrows():
        lin += 1
        ws.cell(row=lin, column=1, value=int(ano)).font = Font(name=FONTE,
                                                               size=9)
        for j, v in enumerate(valores, start=2):
            ws.cell(row=lin, column=j, value=int(v)).font = Font(name=FONTE,
                                                                 size=9)
    fim_t3 = lin

    pct = 100 * len(recorte) / len(dados)
    nota = ws.cell(row=fim_t3 + 1, column=1,
                   value=f"Recorte {ANO_CORTE}–{DATA_REF.year} "
                         f"({pct:.0f}% dos processos).")
    nota.font = Font(name=FONTE, size=8, italic=True, color="595959")

    g3 = LineChart()
    g3.title = (f"Aberturas por ano — 5 substâncias com mais processos "
                f"({ANO_CORTE}–{DATA_REF.year})")
    g3.y_axis.title = "Processos abertos"
    g3.x_axis.title = "Ano do protocolo"
    g3.height, g3.width = 10, 24
    g3.add_data(Reference(ws, min_col=2, max_col=6, min_row=lin0,
                          max_row=fim_t3), titles_from_data=True)
    g3.set_categories(Reference(ws, min_col=1, min_row=lin0 + 1,
                                max_row=fim_t3))
    for serie, cor in zip(g3.series, CORES_SERIES):
        _estilo_serie(serie, cor, linha=True)
    ws.add_chart(g3, f"H{lin0}")  # à direita da tabela (que vai até a col. F)

    # ---- T4: grupo do calcário por ano ------------------------------------
    calc = dados[dados["Substância"].isin(SUBSTANCIAS_CALCARIO)]
    por_ano_calc = (calc["Ano_abertura"].value_counts()
                    .reindex(range(ANO_CORTE, DATA_REF.year + 1),
                             fill_value=0))

    lin0 = fim_t3 + 4
    _cabecalho(ws, lin0, ["Ano", "Processos (grupo do calcário)"])
    lin = lin0
    for ano, qtd in por_ano_calc.items():
        lin += 1
        ws.cell(row=lin, column=1, value=int(ano)).font = Font(name=FONTE,
                                                               size=9)
        ws.cell(row=lin, column=2, value=int(qtd)).font = Font(name=FONTE,
                                                               size=9)
    fim_t4 = lin

    antes = int((calc["Ano_abertura"] < ANO_CORTE).sum())
    nota = ws.cell(
        row=fim_t4 + 1, column=1,
        value=f"Grupo: {', '.join(SUBSTANCIAS_CALCARIO)} — "
              f"{len(calc)} processos no total ({antes} antes de "
              f"{ANO_CORTE}, fora do gráfico).")
    nota.font = Font(name=FONTE, size=8, italic=True, color="595959")

    g4 = BarChart()
    g4.type = "col"
    g4.title = (f"Aberturas por ano — grupo do calcário "
                f"({ANO_CORTE}–{DATA_REF.year})")
    g4.y_axis.title = "Processos abertos"
    g4.x_axis.title = "Ano do protocolo"
    g4.height, g4.width = 9, 24
    g4.add_data(Reference(ws, min_col=2, min_row=lin0, max_row=fim_t4),
                titles_from_data=True)
    g4.set_categories(Reference(ws, min_col=1, min_row=lin0 + 1,
                                max_row=fim_t4))
    _estilo_serie(g4.series[0], COR_CALCARIO)
    g4.legend = None
    g4.gapWidth = 40
    ws.add_chart(g4, f"D{lin0}")

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 16
    for col in "CDEF":
        ws.column_dimensions[col].width = 12

    return {"por_ano": por_ano, "top10": top10, "outras": outras,
            "tabela_top5": tabela, "por_ano_calc": por_ano_calc,
            "linha_final": fim_t4}


def gravar_tabela_bolhas(wb, resumo: pd.DataFrame, caminho_png: Path,
                         linha_inicio: int):
    """Acrescenta a tabela-resumo por substância e o gráfico de bolhas."""
    ws = wb[ABA_GRAFICOS]
    lin0 = linha_inicio
    _cabecalho(ws, lin0, list(resumo.columns))
    lin = lin0
    for linha in resumo.itertuples(index=False):
        lin += 1
        for j, v in enumerate(linha, start=1):
            ws.cell(row=lin, column=j, value=v).font = Font(name=FONTE,
                                                            size=9)
    nota = ws.cell(row=lin + 1, column=1,
                   value="Resumo das 45 substâncias do gráfico de bolhas ao "
                         "lado (todas as substâncias × todos os anos; a "
                         "classificação em famílias é nossa, não da ANM).")
    nota.font = Font(name=FONTE, size=8, italic=True, color="595959")

    from openpyxl.drawing.image import Image as XLImage
    img = XLImage(str(caminho_png))
    fator = 760 / img.width
    img.width = int(img.width * fator)
    img.height = int(img.height * fator)
    ws.add_image(img, f"J{lin0}")  # à direita da tabela (colunas A–G)


# ---------------------------------------------------------------- PNGs


SUPERFICIE, TINTA, MUDO, GRADE = "#fcfcfb", "#0b0b0b", "#898781", "#e1e0d9"


def _preparar_matplotlib():
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    PASTA_GRAFICOS.mkdir(exist_ok=True)
    plt.rcParams.update({
        "font.family": "sans-serif", "font.size": 10,
        "figure.facecolor": SUPERFICIE, "axes.facecolor": SUPERFICIE,
        "axes.edgecolor": GRADE, "axes.labelcolor": MUDO,
        "xtick.color": MUDO, "ytick.color": MUDO,
        "axes.grid": True, "grid.color": GRADE, "grid.linewidth": 0.6,
        "axes.spines.top": False, "axes.spines.right": False,
        "axes.titlecolor": TINTA, "axes.titlesize": 12,
        "axes.titleweight": "bold",
    })
    return plt


def _salvar(plt, fig, nome):
    caminho = PASTA_GRAFICOS / nome
    fig.savefig(caminho, dpi=150, bbox_inches="tight")
    plt.close(fig)
    print(f"  PNG: {caminho.relative_to(RAIZ)}")
    return caminho


def gerar_png_bolhas(dados: pd.DataFrame) -> Path:
    """Bolhas minimalistas: todas as substâncias × todos os anos.

    Sem eixos, grade ou marcações — só rótulos flutuando. Cada família
    tem uma cor; bolha maior = mais processos abertos naquele ano.
    """
    plt = _preparar_matplotlib()
    blocos = blocos_familias(dados)
    ESCALA = 2.2          # área da bolha (pt²) por processo
    X_ROTULO = 1937.0     # coluna dos rótulos (à esquerda das bolhas)

    # Posições verticais: cabeçalho da família, linhas, respiro entre blocos
    linhas, cabecalhos = [], []
    y = 0.0
    for nome_fam, cor, itens in blocos:
        cabecalhos.append((y, nome_fam, cor))
        y -= 1.2
        for substancia, serie in itens:
            linhas.append((y, substancia, cor, serie))
            y -= 1.0
        y -= 1.1
    y_fim = y

    fig, ax = plt.subplots(figsize=(13, 0.26 * (-y_fim + 12)))
    ax.set_axis_off()

    for yy, nome_fam, cor in cabecalhos:
        ax.text(X_ROTULO, yy, nome_fam, ha="right", va="center",
                fontsize=9, fontweight="bold", color=cor)
    for yy, substancia, cor, serie in linhas:
        com_valor = serie[serie > 0]
        # piso de 6 pt²: sem ele, os anos com 1 processo somem do gráfico
        ax.scatter(com_valor.index, [yy] * len(com_valor),
                   s=(com_valor.values * ESCALA).clip(min=6), color=cor,
                   edgecolors=SUPERFICIE, linewidths=0.8, zorder=3)
        ax.text(X_ROTULO, yy, substancia, ha="right", va="center",
                fontsize=8, color="#52514e")

    # Décadas no alto e no pé (referência discreta, sem linhas)
    for dec in range(1940, 2030, 10):
        for yy in (1.4, y_fim - 0.6):
            ax.text(dec, yy, str(dec), ha="center", va="center",
                    fontsize=9, color=MUDO)

    # Título e subtítulo
    ax.text(1912, 4.6, "Todos os minérios de MS ao longo do tempo",
            fontsize=16, fontweight="bold", color=TINTA)
    ax.text(1912, 3.1,
            "Cada bolha: processos minerários abertos na ANM naquele ano "
            f"(SIGMINE, {len(dados)} processos, 1940–{DATA_REF.year}). "
            "Cores por família de substância.",
            fontsize=9, color=MUDO)

    # Legenda de tamanho
    y_leg = y_fim - 3.4
    ax.text(1984, y_leg, "processos no ano:", ha="right", va="center",
            fontsize=8.5, color=MUDO)
    for x, v in ((1990, 5), (1998, 25), (2008, 100)):
        ax.scatter([x], [y_leg], s=v * ESCALA, color="#c3c2b7",
                   edgecolors=SUPERFICIE)
        ax.text(x, y_leg - 1.7, str(v), ha="center", va="center",
                fontsize=8, color=MUDO)

    ax.set_xlim(1911, 2029)
    ax.set_ylim(y_leg - 3.0, 5.6)
    return _salvar(plt, fig, "bolhas_minerios_tempo.png")


def _familia_de(substancia):
    for nome_fam, cor, subs in FAMILIAS:
        if substancia in subs:
            return nome_fam, cor
    return "NÃO CLASSIFICADAS", "#c3c2b7"


def gerar_png_area_acumulada(dados: pd.DataFrame) -> Path:
    """Área empilhada: hectares ACUMULADOS sob processo, por família.

    O 'avanço do território de direito': cada faixa é uma família de
    substâncias; a altura total é quanto chão de MS já foi requisitado.
    """
    plt = _preparar_matplotlib()

    fam = dados["Substância"].map(lambda s: _familia_de(s)[0])
    anos = range(int(dados["Ano_abertura"].min()), DATA_REF.year + 1)
    acumulado = (dados.assign(Família=fam)
                 .pivot_table(index="Ano_abertura", columns="Família",
                              values="Área (ha)", aggfunc="sum",
                              fill_value=0)
                 .reindex(anos, fill_value=0).cumsum())
    # maiores embaixo (empilhamento mais estável de ler)
    ordem = acumulado.iloc[-1].sort_values(ascending=False).index
    acumulado = acumulado[ordem]
    cores = {nome: cor for nome, cor, _ in FAMILIAS}
    cores["NÃO CLASSIFICADAS"] = "#c3c2b7"

    fig, ax = plt.subplots(figsize=(12, 6.5))
    ax.stackplot(acumulado.index, acumulado.T.values,
                 colors=[cores[c] for c in acumulado.columns],
                 linewidth=0.6, edgecolor=SUPERFICIE)
    ax.set_axis_off()

    # Marcas discretas de % do estado, à direita
    for pct in (1, 2, 3, 4, 5):
        y = MS_HA * pct / 100
        ax.plot([DATA_REF.year + 0.3, DATA_REF.year + 1.0], [y, y],
                color=MUDO, linewidth=0.8, clip_on=False)
        ax.text(DATA_REF.year + 1.6, y, f"{pct}% de MS", va="center",
                fontsize=8.5, color=MUDO)

    # Décadas no pé
    for dec in range(1940, 2030, 10):
        ax.text(dec, -MS_HA * 0.0045, str(dec), ha="center", va="top",
                fontsize=9, color=MUDO)

    # Legenda-lista no espaço vazio à esquerda, na ordem do empilhamento
    # (faixa de cima primeiro), com o total final de cada família
    def _fmt_ha(v):
        if v >= 1e6:
            return f"{v/1e6:.2f}".replace(".", ",") + " mi ha"
        return f"{v/1e3:,.0f}".replace(",", ".") + " mil ha"

    y_lista, passo = MS_HA * 0.052, MS_HA * 0.0042
    for i, coluna in enumerate(reversed(list(acumulado.columns))):
        y = y_lista - i * passo
        ax.scatter([1941.5], [y], s=46, color=cores[coluna],
                   edgecolors="none", clip_on=False)
        ax.text(1943.5, y,
                f"{coluna.capitalize()} — "
                f"{_fmt_ha(acumulado[coluna].iloc[-1])}",
                va="center", fontsize=9, color="#52514e")

    total = dados["Área (ha)"].sum()
    ax.text(1939, MS_HA * 0.068,
            "O avanço do território de direito",
            fontsize=16, fontweight="bold", color=TINTA)
    ax.text(1939, MS_HA * 0.0625,
            f"Hectares acumulados sob processo minerário em MS, por família "
            f"de substância (SIGMINE/ANM). Hoje: "
            f"{total/1e6:.1f}".replace(".", ",") + " milhões de ha = "
            f"{100*total/MS_HA:.1f}".replace(".", ",") + "% do estado.",
            fontsize=9, color=MUDO)
    ax.set_xlim(1938, 2037)
    ax.set_ylim(0, MS_HA * 0.072)
    return _salvar(plt, fig, "area_acumulada_familias.png")


def gerar_png_mapa_calor(dados: pd.DataFrame) -> Path:
    """Mapa de calor: substância × ano, cada família na sua cor.

    A tonalidade identifica a família (como no gráfico de bolhas) e a
    intensidade (claro → escuro) é o nº de processos abertos no ano,
    com a MESMA escala em todas as cores.
    """
    import numpy as np
    from matplotlib.colors import LinearSegmentedColormap, PowerNorm, to_rgb
    plt = _preparar_matplotlib()

    blocos = blocos_familias(dados)
    anos = list(range(int(dados["Ano_abertura"].min()), DATA_REF.year + 1))
    bordas_x = np.array(anos + [DATA_REF.year + 1]) - 0.5
    maximo = max(s.max() for _, _, itens in blocos for _, s in itens)
    norma = PowerNorm(0.45, vmin=1, vmax=maximo)

    def rampa(cor_base):
        """Do quase-branco à cor escura, passando pela cor da família."""
        base = np.array(to_rgb(cor_base))
        clara = 0.85 * np.array(to_rgb(SUPERFICIE)) + 0.15 * base
        escura = 0.45 * base
        mapa = LinearSegmentedColormap.from_list(
            "r", [clara, tuple(base), tuple(escura)])
        mapa.set_bad(SUPERFICIE)
        return mapa

    # Cada família vira um bloco de linhas, desenhado com a sua rampa
    rotulos, blocos_desenho = [], []   # (linha_inicial, matriz, rampa)
    linha = 0
    for nome_fam, cor, itens in blocos:
        rotulos.append((linha, nome_fam, cor, True))
        linha += 1
        matriz_fam = []
        for substancia, serie in itens:
            rotulos.append((linha + len(matriz_fam), substancia,
                            "#52514e", False))
            valores = serie.reindex(anos, fill_value=0).astype(float)
            matriz_fam.append(valores.where(valores > 0).tolist())
        blocos_desenho.append((linha, np.ma.masked_invalid(matriz_fam),
                               rampa(cor)))
        linha += len(matriz_fam) + 1
    n = linha - 1

    fig, ax = plt.subplots(figsize=(13, 0.235 * (n + 10)))
    for linha0, matriz_fam, mapa in blocos_desenho:
        ax.pcolormesh(bordas_x,
                      np.arange(linha0, linha0 + matriz_fam.shape[0] + 1)
                      - 0.5,
                      matriz_fam, cmap=mapa, norm=norma,
                      edgecolors=SUPERFICIE, linewidth=0.4)
    ax.invert_yaxis()
    ax.set_axis_off()

    for indice, texto, cor, eh_fam in rotulos:
        ax.text(anos[0] - 2.5, indice, texto, ha="right", va="center",
                fontsize=8.5 if eh_fam else 8,
                fontweight="bold" if eh_fam else "normal", color=cor)
    for dec in range(1940, 2030, 10):
        ax.text(dec, -2.2, str(dec), ha="center", va="center",
                fontsize=9, color=MUDO)

    ax.text(1912, -5.6, "Mapa de calor da corrida mineral em MS",
            fontsize=16, fontweight="bold", color=TINTA)
    ax.text(1912, -3.9,
            "Processos minerários abertos por ano e substância "
            f"(SIGMINE/ANM, {len(dados)} processos). A cor identifica a "
            "família; quanto mais escuro, mais processos no ano.",
            fontsize=9, color=MUDO)

    # Escala de intensidade (igual para todas as famílias), em cinza
    exemplos, x0, y0 = [1, 5, 25, 100, maximo], 1990, n + 3.2
    cinza = rampa("#898781")
    for i, v in enumerate(exemplos):
        x = x0 + i * 4
        ax.add_patch(plt.Rectangle((x - 1.2, y0 - 0.5), 2.4, 1.4,
                                   facecolor=cinza(norma(v)),
                                   edgecolor=SUPERFICIE, clip_on=False))
        ax.text(x, y0 + 2.2, str(int(v)), ha="center", va="center",
                fontsize=8, color=MUDO)
    ax.text(x0 - 4, y0 + 0.2, "processos no ano\n(escala igual em todas "
            "as cores):", ha="right", va="center", fontsize=8.5, color=MUDO)

    ax.set_xlim(1911, 2029)
    ax.set_ylim(y0 + 3.6, -7.0)
    return _salvar(plt, fig, "mapa_calor_minerios.png")


def gerar_png_dormencia(dados: pd.DataFrame) -> Path:
    """Dispersão da dormência: quando o processo nasceu × há quanto tempo
    não se movimenta, colorido pela situação atual (fase agrupada)."""
    plt = _preparar_matplotlib()

    def situacao(fase):
        f = (fase or "").upper()
        if "DISPONIBILIDADE" in f:
            return "Devolvido / em disponibilidade"
        if any(p in f for p in ("PESQUISA", "REQUERER")):
            return "Pesquisa (no papel)"
        if "NÃO CADASTRADO" in f:
            return "Não cadastrado"
        return "Lavra / extração autorizada ou pedida"

    CORES_SITUACAO = {
        "Pesquisa (no papel)": "#2a78d6",
        "Lavra / extração autorizada ou pedida": "#eb6834",
        "Devolvido / em disponibilidade": "#c3c2b7",
        "Não cadastrado": "#c3c2b7",
    }
    validos = dados.dropna(subset=["Anos_sem_movimentacao"])

    fig, ax = plt.subplots(figsize=(11, 6.5))
    for nome, grupo in validos.groupby(validos["Fase"].map(situacao)):
        ax.scatter(grupo["Ano_abertura"], grupo["Anos_sem_movimentacao"],
                   s=10, color=CORES_SITUACAO[nome], alpha=0.4,
                   edgecolors="none", zorder=3, label=nome)

    ax.grid(visible=False)
    for lado in ("left", "bottom"):
        ax.spines[lado].set_visible(False)
    ax.tick_params(length=0, labelsize=9)
    ax.set_xlabel("Ano de abertura do processo", fontsize=9.5)
    ax.set_ylabel("Anos desde a última movimentação", fontsize=9.5)

    ax.set_title("Processos vivos, processos fósseis", loc="left", pad=32)
    ax.text(0, 1.045,
            "Cada ponto é um processo minerário de MS: quando nasceu × há "
            "quanto tempo não se movimenta na ANM (SIGMINE, "
            f"ref. {DATA_REF.strftime('%m/%Y')}).",
            transform=ax.transAxes, fontsize=9, color=MUDO)

    parados = int((validos["Anos_sem_movimentacao"] >= 10).sum())
    ax.annotate(f"{parados} processos sem nenhuma movimentação há 10+ anos",
                xy=(1994, 10), fontsize=8.5, color="#52514e",
                xytext=(1948, 22),
                arrowprops=dict(arrowstyle="-", color=MUDO, linewidth=0.8))
    legenda = ax.legend(loc="upper right", frameon=False, fontsize=8.5,
                        markerscale=2.2, labelcolor=TINTA)
    for alca in legenda.legend_handles:
        alca.set_alpha(1)
    return _salvar(plt, fig, "dormencia_processos.png")


def gerar_pngs(agregados, dados):
    plt = _preparar_matplotlib()

    def salvar(fig, nome):
        _salvar(plt, fig, nome)

    # 1 — série temporal geral
    por_ano = agregados["por_ano"]
    fig, ax = plt.subplots(figsize=(10, 4.5))
    ax.plot(por_ano.index, por_ano.values, color="#256abf", linewidth=2)
    ax.set_title("Processos minerários abertos por ano — MS (SIGMINE/ANM)",
                 loc="left")
    ax.set_xlabel("Ano do protocolo")
    ax.set_ylabel("Processos abertos")
    ax.grid(axis="x", visible=False)
    salvar(fig, "serie_temporal_geral.png")

    # 2 — barras por substância
    top10, outras = agregados["top10"], agregados["outras"]
    nomes = list(top10.index) + ["OUTRAS"]
    valores = list(top10.values) + [outras]
    fig, ax = plt.subplots(figsize=(9, 5))
    y = range(len(nomes))[::-1]
    ax.barh(y, valores, color="#2a78d6", height=0.62)
    ax.barh(y[-1], valores[-1], color="#c3c2b7", height=0.62)
    ax.set_yticks(y, nomes)
    for yi, v in zip(y, valores):
        ax.text(v + 12, yi, f"{v:,}".replace(",", "."),
                va="center", color=TINTA, fontsize=9)
    ax.set_title("Processos por substância — MS (top 10 + outras)",
                 loc="left")
    ax.set_xlabel("Processos")
    ax.grid(axis="y", visible=False)
    salvar(fig, "processos_por_substancia.png")

    # 3 — série por substância (top 5)
    tabela = agregados["tabela_top5"]
    cores = ["#2a78d6", "#eb6834", "#1baf7a", "#eda100", "#e87ba4"]
    fig, ax = plt.subplots(figsize=(10, 5))
    for coluna, cor in zip(tabela.columns, cores):
        ax.plot(tabela.index, tabela[coluna], color=cor, linewidth=2,
                label=coluna)
        ax.annotate(coluna, xy=(tabela.index[-1], tabela[coluna].iloc[-1]),
                    xytext=(6, 0), textcoords="offset points",
                    va="center", fontsize=8.5, color=TINTA)
    ax.set_title("Aberturas por ano — 5 substâncias com mais processos",
                 loc="left")
    ax.set_xlabel("Ano do protocolo")
    ax.set_ylabel("Processos abertos")
    ax.legend(loc="upper left", frameon=False, fontsize=8.5,
              labelcolor=TINTA)
    ax.grid(axis="x", visible=False)
    ax.set_xlim(tabela.index[0], tabela.index[-1] + 4)
    salvar(fig, "serie_por_substancia_top5.png")

    # 4 — grupo do calcário
    por_ano_calc = agregados["por_ano_calc"]
    fig, ax = plt.subplots(figsize=(10, 4.5))
    ax.bar(por_ano_calc.index, por_ano_calc.values, color="#008300",
           width=0.7)
    ax.set_title("Aberturas por ano — grupo do calcário "
                 "(calcário, calcítico, dolomítico e calcita)", loc="left")
    ax.set_xlabel("Ano do protocolo")
    ax.set_ylabel("Processos abertos")
    ax.grid(axis="x", visible=False)
    salvar(fig, "serie_calcario.png")


# ---------------------------------------------------------------- principal


def principal():
    entrada = Path(sys.argv[1]) if len(sys.argv) > 1 else ENTRADA_PADRAO
    if not entrada.exists():
        raise SystemExit(f"Arquivo não encontrado: {entrada}")

    print(f"Lendo {entrada.name} …")
    df = pd.read_excel(entrada, sheet_name=ABA_DADOS, header=3)
    print(f"  {len(df)} processos na aba '{ABA_DADOS}'.")

    dados = extrair_temporal(df)
    sem_data = int(dados["Data_ultimo_evento"].isna().sum())
    print(f"  Ano de abertura extraído em 100% das linhas; "
          f"data do último evento ausente em {sem_data} linha(s).")

    wb = load_workbook(entrada)
    gravar_aba_dados(wb, dados)
    agregados = gravar_aba_graficos(wb, dados)

    caminho_bolhas = gerar_png_bolhas(dados)
    gravar_tabela_bolhas(wb, resumo_bolhas(dados), caminho_bolhas,
                         agregados["linha_final"] + 3)

    wb.save(entrada)
    print(f"Abas '{ABA_NOVA}' e '{ABA_GRAFICOS}' gravadas em {entrada.name}.")

    gerar_pngs(agregados, dados)
    gerar_png_area_acumulada(dados)
    gerar_png_mapa_calor(dados)
    gerar_png_dormencia(dados)
    print("Concluído.")


if __name__ == "__main__":
    principal()
